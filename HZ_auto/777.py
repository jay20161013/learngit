import pyautogui
import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By

# 读取Excel表格数据

pyautogui.FAILSAFE = False  # 防错

wb = openpyxl.load_workbook(r'.\cs.xlsx')
sheet = wb.get_sheet_by_name('报缴用表')
iedriver = R".\IEDriverServer.exe"
browser = webdriver.Ie(iedriver)
browser.get('http://192.168.81.249/friends/login.jsp')
# browser.find_element_by_xpath('/html/body/div[1]/div/div[4]/span/a[1]').click()
# time.sleep(1)

# 输入账号密码 点击登录
browser.find_element_by_xpath('//*[@name="user"]').send_keys('lh')
browser.find_element_by_xpath('//*[@name="password"]').send_keys('414871250@')
browser.find_element_by_xpath('//*[@name="imageField"]').click()

for handle in browser.window_handles:
    browser.switch_to.window(handle)

# browser.close()

browser.get('http://192.168.81.249/friends/buy/buy_in_add.vw')
browser.find_element_by_xpath('//*[@msg="备注"]').send_keys(sheet.cell(2, 13).value)
browser.find_element_by_xpath('//*[@id="v_storeid"]').send_keys('上海采购过渡仓库')
pyautogui.typewrite(['enter'], '0.25')
browser.find_element_by_xpath('//*[@id="dwmc"]').send_keys('上海南方')
pyautogui.typewrite(['enter'], '0.25')



# pyautogui.moveTo(1130, 45, duration=0.11)12

# pyautogui.click(clicks=(sheet.max_row - 1) * 2, interval=0.05)
# pyautogui.moveTo(842, 299)
sc=pyautogui.screenshot()
number7_location = pyautogui.locateOnScreen('tjyh.png')#传入按钮的图片
pyautogui.click(pyautogui.center(number7_location ),clicks=(sheet.max_row - 1) * 2, interval=0.05)

x1,y1=pyautogui.center(pyautogui.locateOnScreen('cpbm.png'))
x2,y2=pyautogui.center(pyautogui.locateOnScreen('yjdhl.png'))
print(sheet.max_row + 1)
for i in range(1, sheet.max_row + 1):
    pyautogui.moveTo(x1, (y1 + 18.5 * i ))
    pyautogui.click()
    print(i,str(sheet.cell(i + 1, 5).value))
    pyautogui.typewrite(str(sheet.cell(i + 1, 5).value))
    pyautogui.typewrite(['enter'], '0.25')
    time.sleep(3)
    pyautogui.moveRel(250, 0)   #相对移动   moveto 绝对移动
    pyautogui.click()
    pyautogui.typewrite(str(sheet.cell(i + 1, 6).value))
    #pyautogui.typewrite(['enter'], '0.25')
    time.sleep(1)

