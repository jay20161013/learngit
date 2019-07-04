import cx_Oracle
import openpyxl

wb = openpyxl.load_workbook(r'.\cs.xlsx')
if 'cs123'in wb.get_sheet_names():
    wb.remove_sheet('cs123')
    print(wb.get_sheet_names())
ws = wb.create_sheet('cs123')
ws.append(['TIME', 'TITLE', 'A-Z'])

conn = cx_Oracle.connect('nanfang/nanfangerp!@#@192.168.81.249:1521/NANFANG.ACT')

cursor = conn.cursor()
# cursor.execute("SELECT * FROM emp")
# rows = cursor.fetchall()  # 得到所有数据集
#
# ws.append(cursor.fetchall())
# wb.save(r'.\cs.xlsx')
# for row in rows:
#     print("%d, %s, %s, %s" % (row[0], row[1], row[2], row[4]))
#
# print("Number of rows returned: %d" % cursor.rowcount)

cursor.execute("SELECT * FROM emp")
while (True):
    row = cursor.fetchone()  # 逐行得到数据集
    if row == None:
        break
#     #print("%d, %s, %s, %s" % (row[0], row[1], row[2], row[4]))
#     time=row[0]
#     title=row[1]
#     a_z=row[2]
#     xx=row[4]
#     ws.append([time,title,a_z,xx])
#     ws.cell(2,5).value=row[4]
for i in range(400):
    row = cursor.fetchone()  # 逐行得到数据集
    TIME = row[0]
    TITLE = row[1]
    A_Z = row[4]
    ws.append([TIME, TITLE, A_Z])
#print("Number of rows returned: %d" % cursor.rowcount)
wb.save(r'.\cs.xlsx')
cursor.close()
conn.close()


