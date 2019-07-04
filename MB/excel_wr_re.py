import openpyxl
from openpyxl.utils import get_column_letter
from random import choice
from time import time
import datetime
wb = openpyxl.load_workbook(r'.\cs.xlsx')
ws=wb.create_sheet('cs123')
ws.append(['TIME', 'TITLE', 'A-Z'])

# 输入内容（500行数据）
for i in range(500):
    TIME = datetime.datetime.now().strftime("%H:%M:%S")
    TITLE = str(time())
    A_Z = get_column_letter(choice(range(1, 50)))
    ws.append([TIME, TITLE, A_Z])
print(ws.max_row)
ws.cell(2,5).value=11

wb.save(r'.\cs.xlsx')

